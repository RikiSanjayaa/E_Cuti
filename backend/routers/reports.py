from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import FileResponse, StreamingResponse
from sqlalchemy.orm import Session, joinedload
from datetime import date, timedelta
import pandas as pd
from io import BytesIO
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, KeepTogether
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, KeepTogether
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from backend.core import database, auth
from backend import models, schemas

router = APIRouter(
    prefix="/api/reports",
    tags=["Reports"]
)

@router.get("/summary", response_model=schemas.AnalyticsSummary)
async def get_analytics_summary(
    start_date: date = Query(None),
    end_date: date = Query(None),
    leave_type: str = Query(None),
    personnel_id: int = Query(None),
    current_user: models.User = Depends(auth.get_current_user),
    db: Session = Depends(database.get_db)
):
    query = db.query(models.LeaveHistory)\
        .join(models.Personnel)\
        .join(models.LeaveType)\
        .options(joinedload(models.LeaveHistory.personnel), joinedload(models.LeaveHistory.leave_type))
    
    if start_date:
        query = query.filter(models.LeaveHistory.tanggal_mulai >= start_date)
    if end_date:
        query = query.filter(models.LeaveHistory.tanggal_mulai <= end_date)

    if leave_type and leave_type != 'all':
        query = query.filter(models.LeaveType.code == leave_type)
        
    if personnel_id:
        query = query.filter(models.LeaveHistory.personnel_id == personnel_id)
        
    leaves = query.all()
    
    total_days = sum(l.jumlah_hari for l in leaves)
    unique_personel = len(set(l.personnel_id for l in leaves))
    
    return {
        "total_records": len(leaves),
        "total_days": total_days,
        "unique_personel": unique_personel,
        "data": leaves
    }

@router.get("/export")
async def export_report(
    format: str = Query(..., pattern="^(pdf|excel)$"),
    start_date: date = Query(None),
    end_date: date = Query(None),
    leave_type: str = Query(None),
    personnel_id: int = Query(None),
    month: int = Query(None, ge=1, le=12),
    year: int = Query(None, ge=2000),
    current_user: models.User = Depends(auth.get_current_user_from_token),
    db: Session = Depends(database.get_db)
):
    # Filter data
    query = db.query(models.LeaveHistory)\
        .join(models.Personnel)\
        .join(models.LeaveType)\
        .options(joinedload(models.LeaveHistory.leave_type))
    
    # Date Filtering Logic
    if start_date:
        query = query.filter(models.LeaveHistory.tanggal_mulai >= start_date)
    if end_date:
        query = query.filter(models.LeaveHistory.tanggal_mulai <= end_date)
        
    # Fallback to Month/Year if provided and no start/end date
    if not start_date and not end_date:
        if month and year:
            from sqlalchemy import extract
            query = query.filter(extract('month', models.LeaveHistory.tanggal_mulai) == month)
            query = query.filter(extract('year', models.LeaveHistory.tanggal_mulai) == year)
        elif year:
            from sqlalchemy import extract
            query = query.filter(extract('year', models.LeaveHistory.tanggal_mulai) == year)
            
    # Other filters
    if leave_type and leave_type != 'all':
        query = query.filter(models.LeaveType.code == leave_type)
        
    if personnel_id:
        query = query.filter(models.LeaveHistory.personnel_id == personnel_id)
        
    leaves = query.all()
    
    # Prepare data for report
    data_list = []
    
    for idx, leave in enumerate(leaves, 1):
        # Get leave type name from relationship
        jenis_izin_str = leave.leave_type.name if leave.leave_type else "-"
        
        # Format Date
        tgl_str = leave.tanggal_mulai.strftime("%d %b %Y") if leave.tanggal_mulai else "-"
        
        # Combine for cleaner output
        jenis_izin_display = f"{jenis_izin_str}\n(Mulai: {tgl_str})\n({leave.jumlah_hari} hari)"

        data_list.append({
            "NO": idx,
            "NAMA": leave.personnel.nama,
            "PANGKAT": leave.personnel.pangkat,
            "NRP/NIP": leave.personnel.nrp,
            "JABATAN": leave.personnel.jabatan,
            "JENIS IZIN": jenis_izin_display,
            "KETERANGAN": leave.alasan
        })
        
    if not data_list:
        raise HTTPException(status_code=404, detail="No data found for the selected period")

    if format == "excel":
        # Prepare granular data for Excel
        excel_data = []
        for idx, leave in enumerate(leaves, 1):
            # Calculate end date
            tgl_mulai = leave.tanggal_mulai
            tgl_selesai = tgl_mulai + timedelta(days=leave.jumlah_hari - 1) if tgl_mulai else None
            
            # Format Dates
            tgl_mulai_str = tgl_mulai.strftime("%d-%m-%Y") if tgl_mulai else "-"
            tgl_selesai_str = tgl_selesai.strftime("%d-%m-%Y") if tgl_selesai else "-"
            
            excel_data.append({
                "NO": idx,
                "NAMA": leave.personnel.nama,
                "PANGKAT": leave.personnel.pangkat,
                "NRP/NIP": leave.personnel.nrp,
                "JABATAN": leave.personnel.jabatan,
                "JENIS CUTI": leave.leave_type.name if leave.leave_type else "-",
                "TANGGAL MULAI": tgl_mulai_str,
                "TANGGAL SELESAI": tgl_selesai_str,
                "DURASI (Hari)": leave.jumlah_hari,
                "KETERANGAN": leave.alasan or "-"
            })

        df = pd.DataFrame(excel_data)
        stream = BytesIO()
        
        # Use openpyxl engine with styling
        with pd.ExcelWriter(stream, engine='openpyxl') as writer:
            # Write data starting from row 4 (to leave space for header)
            df.to_excel(writer, index=False, sheet_name='Laporan Cuti', startrow=3)
            
            worksheet = writer.sheets['Laporan Cuti']
            
            # Styles
            header_font = Font(name='Calibri', size=11, bold=True)
            title_font = Font(name='Calibri', size=14, bold=True)
            subtitle_font = Font(name='Calibri', size=12, bold=True)
            
            center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            thin_border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )
            
            header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid") # Light Gray
            
            # 1. Add Title & Period
            # Merge rows 1 and 2 for Title & Subtitle
            last_col_letter = get_column_letter(len(df.columns))
            
            # Title
            worksheet.merge_cells(f'A1:{last_col_letter}1')
            title_cell = worksheet['A1']
            title_cell.value = "LAPORAN DATA IZIN/CUTI PERSONEL"
            title_cell.font = title_font
            title_cell.alignment = center_align
            
            # Period
            worksheet.merge_cells(f'A2:{last_col_letter}2')
            subtitle_cell = worksheet['A2']
            
            if start_date and end_date:
                period_str = f"PERIODE: {start_date.strftime('%d-%m-%Y')} s/d {end_date.strftime('%d-%m-%Y')}"
            elif month and year:
                 period_str = f"PERIODE: {month}/{year}"
            elif year:
                period_str = f"TAHUN {year}"
            else:
                period_str = "SEMUA DATA"
                
            subtitle_cell.value = period_str
            subtitle_cell.font = subtitle_font
            subtitle_cell.alignment = center_align

            # 2. Format Table Header (Row 4)
            for col_num, column_title in enumerate(df.columns, 1):
                col_letter = get_column_letter(col_num)
                cell = worksheet[f"{col_letter}4"]
                cell.font = header_font
                cell.alignment = center_align
                cell.fill = header_fill
                cell.border = thin_border
                
            # 3. Format Data Rows & Column Widths
            # Set Column Widths
            col_widths = {
                'A': 5,   # NO
                'B': 30,  # NAMA
                'C': 15,  # PANGKAT
                'D': 15,  # NRP
                'E': 25,  # JABATAN
                'F': 20,  # JENIS CUTI
                'G': 15,  # TGL MULAI
                'H': 15,  # TGL SELESAI
                'I': 10,  # DURASI
                'J': 30   # KETERANGAN
            }
            
            for col_letter, width in col_widths.items():
                if col_letter <= last_col_letter:
                    worksheet.column_dimensions[col_letter].width = width

            # Apply borders and alignment to all data cells
            last_row = 4
            for row in range(5, 5 + len(df)):
                last_row = row
                for col in range(1, len(df.columns) + 1): 
                    col_letter = get_column_letter(col)
                    cell = worksheet[f"{col_letter}{row}"]
                    cell.border = thin_border
                    
                    # Center align for NO, Dates, Durasi 
                    # Columns: A(1) NO, G(7) Start, H(8) End, I(9) Durasi
                    if col in [1, 7, 8, 9]: 
                        cell.alignment = center_align
                    else:
                        cell.alignment = left_align

            # 4. Footer: Total Personnel
            footer_row = last_row + 1
            # Merge all columns for label
            worksheet.merge_cells(f'A{footer_row}:{last_col_letter}{footer_row}')
            footer_label_cell = worksheet[f'A{footer_row}']
            footer_label_cell.value = f"Jumlah Personel yang Ijin: {len(df)} Orang"
            footer_label_cell.font = header_font
            footer_label_cell.alignment = Alignment(horizontal='right', vertical='center')
            
            # User requested "just text, no box/border"
            # So we do NOT apply border to the footer cell using thin_border
            # footer_label_cell.border = thin_border

            # 5. Signature Block
            sig_start_row = footer_row + 3
            sig_col_start = len(df.columns) - 2 # Start signature a bit to the right
            
            # Date Place
            today_str = date.today().strftime("%d %B %Y")
            place_cell = worksheet.cell(row=sig_start_row, column=sig_col_start)
            place_cell.value = f"Mataram, {today_str}"
            place_cell.alignment = Alignment(horizontal='center')
            
            # Role
            role_cell_1 = worksheet.cell(row=sig_start_row + 1, column=sig_col_start)
            role_cell_1.value = "A.N. KARO LOGISTIK POLDA NTB"
            role_cell_1.alignment = Alignment(horizontal='center')
            
            role_cell_2 = worksheet.cell(row=sig_start_row + 2, column=sig_col_start)
            role_cell_2.value = "KASUBAGRENMIN"
            role_cell_2.alignment = Alignment(horizontal='center')
            
            # Name
            name_cell = worksheet.cell(row=sig_start_row + 7, column=sig_col_start)
            name_cell.value = "ETEK RIAWAN, S.E."
            name_cell.font = Font(bold=True, underline='single')
            name_cell.alignment = Alignment(horizontal='center')
            
            # NRP
            nrp_cell = worksheet.cell(row=sig_start_row + 8, column=sig_col_start)
            nrp_cell.value = "KOMPOL NRP 73120869"
            nrp_cell.alignment = Alignment(horizontal='center')

        stream.seek(0)
        
        filename = f"Laporan_Cuti_{year}_{month if month else 'All'}.xlsx"
        return StreamingResponse(
            stream, 
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    elif format == "pdf":
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer, 
            pagesize=landscape(A4),
            leftMargin=30,
            rightMargin=30,
            topMargin=30, 
            bottomMargin=30
        )
        elements = []
        
        styles = getSampleStyleSheet()
        
        # 1. KOP SURAT (Header)
        header_style = ParagraphStyle(
            'Header',
            parent=styles['Normal'],
            fontName='Times-Bold',
            fontSize=12,
            alignment=1,
            leading=14
        )
        
        elements.append(Paragraph("KEPOLISIAN NEGARA REPUBLIK INDONESIA", header_style))
        elements.append(Paragraph("DAERAH NUSA TENGGARA BARAT", header_style))
        elements.append(Paragraph("<u>RO BIRO LOGISTIK</u>", header_style))
        elements.append(Spacer(1, 10))
        
        # Title
        title_style = ParagraphStyle(
            'Title',
            parent=styles['Heading1'],
            fontName='Times-Bold',
            fontSize=14,
            alignment=1,
            leading=18,
            spaceAfter=20
        )
        
        report_title = f"LAPORAN DATA IZIN/CUTI PERSONEL"
        
        if start_date and end_date:
            period_str = f"PERIODE: {start_date.strftime('%d-%m-%Y')} s/d {end_date.strftime('%d-%m-%Y')}"
        elif month and year:
             period_str = f"PERIODE: {month}/{year}"
        elif year:
            period_str = f"TAHUN {year}"
        else:
            period_str = "SEMUA DATA"
        
        elements.append(Paragraph(report_title, title_style))
        elements.append(Paragraph(period_str, ParagraphStyle('SubTitle', parent=title_style, fontSize=12)))
        elements.append(Spacer(1, 15))
        
        # 2. TABLE DATA
        # Columns: NO, NAMA, PANGKAT, NRP, JABATAN, JENIS CUTI, MULAI, SELESAI, DURASI, KETERANGAN
        headers = ["NO", "NAMA", "PANGKAT", "NRP/NIP", "JABATAN", "JENIS CUTI", "TANGGAL MULAI", "TANGGAL SELESAI", "DURASI (Hari)", "KETERANGAN"]
        
        cell_style = ParagraphStyle(
            'CellStyle',
            parent=styles['Normal'],
            fontName='Times-Roman',
            fontSize=9,
            leading=11
        )
        
        center_cell_style = ParagraphStyle(
            'CenterCellStyle',
            parent=cell_style,
            alignment=1
        )
        
        header_cell_style = ParagraphStyle(
            'HeaderCellStyle',
            parent=styles['Normal'],
            fontName='Times-Bold',
            fontSize=9,
            alignment=1,
            leading=11
        )
        
        table_data = [[Paragraph(h, header_cell_style) for h in headers]]
        
        for idx, item in enumerate(leaves, 1):
            jenis_izin_str = item.leave_type.name if item.leave_type else "-"
            
            # Calculate end date
            tgl_mulai = item.tanggal_mulai
            tgl_selesai = tgl_mulai + timedelta(days=item.jumlah_hari - 1) if tgl_mulai else None
            
            tgl_mulai_str = tgl_mulai.strftime("%d-%m-%Y") if tgl_mulai else "-"
            tgl_selesai_str = tgl_selesai.strftime("%d-%m-%Y") if tgl_selesai else "-"
            
            row = [
                str(idx),
                Paragraph(item.personnel.nama, cell_style),
                Paragraph(item.personnel.pangkat, cell_style),
                Paragraph(str(item.personnel.nrp), center_cell_style),
                Paragraph(item.personnel.jabatan, cell_style),
                Paragraph(jenis_izin_str, cell_style),
                Paragraph(tgl_mulai_str, center_cell_style),
                Paragraph(tgl_selesai_str, center_cell_style),
                Paragraph(str(item.jumlah_hari), center_cell_style),
                Paragraph(item.alasan, cell_style)
            ]
            table_data.append(row)
           
        # Total Width approx 780 (A4 Landscape margins 30 off 842)
        # 30+130+70+70+90+90+65+65+40+130 = 780
        col_widths = [30, 130, 70, 70, 90, 80, 65, 65, 45, 135]
        
        t = Table(table_data, colWidths=col_widths, repeatRows=1)
        
        report_table_style = TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'), # Header Center
            ('ALIGN', (0, 1), (0, -1), 'CENTER'), # NO Center
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ])
        t.setStyle(report_table_style)
        elements.append(t)
        
        # Total Personnel Row (Manual Table Row append didn't quite work for colspan in simple loop easily without span definition)
        # Add a separate small table or spacer + text for summary?
        # Better: Add row to table_data with spans (ReportLab supports spans)
        # But for simplicity, let's append a Paragraph below
        elements.append(Spacer(1, 10))
        
        summary_style = ParagraphStyle(
            'Summary',
            parent=styles['Normal'],
            fontName='Times-Bold',
            fontSize=10,
            alignment=2 # Right align
        )
        elements.append(Paragraph(f"Jumlah Personel yang Ijin: {len(leaves)} Orang", summary_style))
        
        elements.append(Spacer(1, 30))
        
        # 3. SIGNATURE BLOCK
        date_str = date.today().strftime("%d %B %Y")
        
        sig_style = ParagraphStyle(
            'Signature',
            parent=styles['Normal'],
            fontName='Times-Roman',
            fontSize=11,
            alignment=1,
            leading=14
        )
        
        sig_header = f"Mataram, {date_str}<br/>A.N. KARO LOGISTIK POLDA NTB<br/>KASUBAGRENMIN"
        sig_name = f"<br/><br/><br/><br/><br/><u><b>ETEK RIAWAN, S.E.</b></u><br/>KOMPOL NRP 73120869"
        
        sig_data = [[
            "",
            Paragraph(sig_header + sig_name, sig_style)
        ]]
        
        sig_table = Table(sig_data, colWidths=[450, 250])
        sig_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('ALIGN', (1, 0), (1, 0), 'CENTER'),
        ]))
        
        elements.append(KeepTogether(sig_table))
        
        doc.build(elements)
        buffer.seek(0)
        
        filename = f"Laporan_Cuti_{year}_{month if month else 'All'}.pdf"
        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
